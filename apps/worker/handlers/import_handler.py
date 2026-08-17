"""Streaming CSV/XLSX import worker with explicit validation and lineage outcomes."""
from __future__ import annotations

import csv
import io
import os
from dataclasses import replace
from tempfile import SpooledTemporaryFile
from time import perf_counter
from typing import Any, BinaryIO, Iterable, Iterator, Mapping, TypeVar
from uuid import UUID

from openpyxl import load_workbook

from packages.application.import_pipeline.import_service import ImportValidationService
from packages.domain.import_pipeline.entities import ImportRow, ImportRowError
from packages.domain.shared.enums import ImportJobStatus
from packages.infrastructure.db.repositories.import_job import (
    ImportExecutionContext,
    ImportJobRepository,
    PersistedImportRow,
)
from packages.infrastructure.storage.s3 import StoragePort

_SPOOL_MEMORY_BYTES = 8 * 1024 * 1024
_BATCH_SIZE = 500


class UnsupportedImportFileError(ValueError):
    """A source file is neither CSV nor XLSX."""


class ImportWorkerHandler:
    """Processes one validated import without logging unmasked source content."""

    def __init__(self, repository: ImportJobRepository, storage: StoragePort) -> None:
        self._repository = repository
        self._storage = storage
        self.last_duration_seconds = 0.0

    async def validate(self, import_job_id: UUID) -> str:
        """Stream the original file, retain row validation lineage, then write an error report."""
        started = perf_counter()
        context = await self._repository.get_context(import_job_id)
        raw_rows = await self._read_source_rows(context)
        existing_keys = await self._repository.source_record_keys(context.job.source_system)
        mapped_job = replace(context.job, status=ImportJobStatus.MAPPED)
        validation = ImportValidationService().validate(
            job=mapped_job,
            mapping=context.mapping,
            raw_rows=raw_rows,
            existing_source_record_keys=existing_keys,
        )
        error_object_key = await self._upload_validation_report(import_job_id, validation.rows)
        await self._repository.persist_validation(
            import_job_id, validation, error_object_key=error_object_key
        )
        self.last_duration_seconds = perf_counter() - started
        return validation.job.status.value

    async def execute(self, import_job_id: UUID) -> str:
        """Commit valid rows in batches; duplicate source records remain replay-safe."""
        started = perf_counter()
        if not await self._repository.start_execution(import_job_id):
            raise ValueError("Import job must be QUEUED before execution.")
        context = await self._repository.get_context(import_job_id)
        pending = await self._repository.valid_uncommitted_rows(import_job_id)
        for batch in _batches(pending, _BATCH_SIZE):
            await self._repository.commit_rows(context, batch)
        report_rows = await self._repository.error_report_rows(import_job_id)
        error_object_key = await self._upload_error_report(import_job_id, report_rows)
        status = await self._repository.finish_execution(
            import_job_id, error_object_key=error_object_key
        )
        self.last_duration_seconds = perf_counter() - started
        return status

    async def _read_source_rows(self, context: ImportExecutionContext) -> list[dict[str, str]]:
        """Download to a bounded spool, then parse with CSV/XLSX read-only iterators."""
        with SpooledTemporaryFile(max_size=_SPOOL_MEMORY_BYTES, mode="w+b") as source:
            await self._storage.download_fileobj(context.job.object_key, source)
            return list(stream_import_rows(source, context.job.original_filename))

    async def _upload_validation_report(
        self, import_job_id: UUID, rows: Iterable[ImportRow]
    ) -> str | None:
        report_rows = [
            {
                "row_number": row.row_number,
                "source_record_key": row.source_record_key or "",
                "field_name": error.field_name or "",
                "error_code": error.error_code,
                "message": error.message,
                "severity": error.severity,
            }
            for row in rows
            for error in row.errors
        ]
        return await self._upload_error_report(import_job_id, report_rows)

    async def _upload_error_report(
        self, import_job_id: UUID, rows: Iterable[Mapping[str, Any]]
    ) -> str | None:
        fieldnames = (
            "row_number",
            "source_record_key",
            "field_name",
            "error_code",
            "message",
            "severity",
        )
        safe_rows = [{field: row.get(field, "") for field in fieldnames} for row in rows]
        if not safe_rows:
            return None
        object_key = f"imports/{import_job_id}/error_report.csv"
        buffer = io.StringIO(newline="")
        writer = csv.DictWriter(
            buffer,
            fieldnames=fieldnames,
        )
        writer.writeheader()
        writer.writerows(safe_rows)
        payload = io.BytesIO(buffer.getvalue().encode("utf-8"))
        await self._storage.upload_fileobj(object_key, payload, content_type="text/csv; charset=utf-8")
        return object_key


def stream_import_rows(source: BinaryIO, filename: str) -> Iterator[dict[str, str]]:
    """Read row dictionaries incrementally from a CSV or XLSX source."""
    extension = os.path.splitext(filename.lower())[1]
    source.seek(0)
    if extension == ".csv":
        yield from _stream_csv_rows(source)
        return
    if extension == ".xlsx":
        yield from _stream_xlsx_rows(source)
        return
    raise UnsupportedImportFileError("Only .csv and .xlsx files are supported.")


def _stream_csv_rows(source: BinaryIO) -> Iterator[dict[str, str]]:
    text_source = io.TextIOWrapper(source, encoding="utf-8-sig", newline="")
    try:
        for row in csv.DictReader(text_source):
            yield {str(key): "" if value is None else str(value) for key, value in row.items() if key is not None}
    finally:
        text_source.detach()


def _stream_xlsx_rows(source: BinaryIO) -> Iterator[dict[str, str]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        values = worksheet.iter_rows(values_only=True)
        headers = next(values, None)
        if not headers:
            return
        names = [str(value).strip() if value is not None else "" for value in headers]
        if not any(names):
            return
        for values_row in values:
            yield {
                name: "" if value is None else str(value)
                for name, value in zip(names, values_row, strict=False)
                if name
            }
    finally:
        workbook.close()


T = TypeVar("T")


def _batches(items: list[T], size: int) -> Iterator[list[T]]:
    for start in range(0, len(items), size):
        yield items[start : start + size]

