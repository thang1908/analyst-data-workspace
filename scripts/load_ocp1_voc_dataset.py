"""Load and classify the OCP1 synthetic VOC workbook for local dashboard use.

The source workbook contains trusted labels for sentiment, severity, journey and
technical group, but those labels do not use the application's taxonomy codes.
This loader preserves every source label in ``source_metadata_json`` and creates
an accepted, traceable classification using taxonomy release 3.0.1.

It intentionally removes only records created by ``analytics-demo``.  It never
touches taxonomy, configuration, import jobs, or feedback from other sources.

Usage:
    uv run python scripts/load_ocp1_voc_dataset.py \
        --input /path/to/VOC_OCP1_100000_synthetic_unique_Journey2.xlsx
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from collections.abc import Iterator
from typing import Any
from uuid import NAMESPACE_URL, UUID, uuid5

from openpyxl import load_workbook
from sqlalchemy import text

from packages.infrastructure.db.session import AsyncSessionLocal, engine

PROJECT_ID = UUID("00000000-0000-0000-0000-000000000001")
DEMO_SOURCE_SYSTEM = "analytics-demo"
SOURCE_SYSTEM = "ocp1-voc-synthetic"
NAMESPACE = "https://cx-platform.local/ocp1-voc-synthetic"
TAXONOMY_VERSION = "3.0.1"
EXPECTED_HEADERS = (
    "TC ID",
    "Feedback",
    "Sentiment",
    "Nhóm nguyên nhân kỹ thuật",
    "Mức độ nghiêm trọng",
    "Journey",
    "Journey 2",
    "NPS",
    "CSAT",
    "CES",
)


@dataclass(frozen=True)
class SourceRecord:
    record_key: str
    feedback: str
    sentiment: str
    technical_group: str
    severity: str
    journey: str
    journey_2: str
    nps: str | None
    csat: str | None
    ces: str | None


@dataclass(frozen=True)
class ClassificationCodes:
    stage_code: str
    lifecycle_step_code: str
    service_request_step_code: str
    service_code: str
    issue_code: str
    sentiment: str
    severity: str


def stable_id(name: str) -> UUID:
    return uuid5(NAMESPACE_URL, f"{NAMESPACE}/{name}")


def normalize_header(value: Any) -> str:
    return " ".join(str(value or "").split())


def optional_text(value: Any) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def required_text(value: Any, row_number: int, field: str) -> str:
    normalized = optional_text(value)
    if normalized is None:
        raise ValueError(f"Dòng {row_number}: thiếu giá trị bắt buộc '{field}'.")
    return normalized


def load_source_records(path: Path) -> list[SourceRecord]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Sheet" not in workbook.sheetnames:
            raise ValueError("Workbook phải có sheet dữ liệu tên 'Sheet'.")
        sheet = workbook["Sheet"]
        rows = sheet.iter_rows(values_only=True)
        headers = tuple(normalize_header(value) for value in next(rows, ()))
        if headers != EXPECTED_HEADERS:
            raise ValueError(
                "Header Excel không đúng. Cần đúng 10 cột: " + ", ".join(EXPECTED_HEADERS)
            )

        records: list[SourceRecord] = []
        seen_keys: set[str] = set()
        for row_number, row in enumerate(rows, start=2):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            record = SourceRecord(
                record_key=required_text(row[0], row_number, "TC ID"),
                feedback=required_text(row[1], row_number, "Feedback"),
                sentiment=required_text(row[2], row_number, "Sentiment"),
                technical_group=required_text(row[3], row_number, "Nhóm nguyên nhân kỹ thuật"),
                severity=required_text(row[4], row_number, "Mức độ nghiêm trọng"),
                journey=required_text(row[5], row_number, "Journey"),
                journey_2=required_text(row[6], row_number, "Journey 2"),
                nps=optional_text(row[7]),
                csat=optional_text(row[8]),
                ces=optional_text(row[9]),
            )
            if record.record_key in seen_keys:
                raise ValueError(f"Dòng {row_number}: TC ID bị trùng: {record.record_key}.")
            seen_keys.add(record.record_key)
            records.append(record)
    finally:
        workbook.close()

    if not records:
        raise ValueError("Workbook không có phản hồi để nạp.")
    return records


def classify_issue(record: SourceRecord) -> tuple[str, str]:
    """Map the trusted technical group, refined by clear text signals, to taxonomy."""
    feedback = record.feedback.casefold()
    group_defaults = {
        "Bảo vệ": ("SV-08", "IS-08-02"),
        "Điện nước": ("SV-07", "IS-07-01"),
        "PCCC": ("SV-08", "IS-08-03"),
        "Vệ sinh": ("SV-09", "IS-09-01"),
        "Thang máy": ("SV-07", "IS-07-01"),
        "Bãi xe": ("SV-05", "IS-05-02"),
        "App cư dân": ("SV-03", "IS-03-02"),
        "Hồ bơi": ("SV-06", "IS-06-01"),
        "Internet": ("SV-07", "IS-07-01"),
        "Gym": ("SV-06", "IS-06-01"),
    }
    try:
        service_code, issue_code = group_defaults[record.technical_group]
    except KeyError as error:
        raise ValueError(
            f"TC ID {record.record_key}: nhóm kỹ thuật không hỗ trợ: {record.technical_group}."
        ) from error

    if record.technical_group == "Bảo vệ" and any(
        signal in feedback for signal in ("trộm", "mất cắp", "xâm nhập", "người lạ")
    ):
        return "SV-08", "IS-08-01"
    if record.technical_group in {"Điện nước", "Thang máy"} and any(
        signal in feedback
        for signal in ("rò", "ngập", "chập", "điện giật", "mùi khét", "kẹt", "mắc kẹt")
    ):
        return "SV-07", "IS-07-02"
    if record.technical_group == "Vệ sinh":
        if any(signal in feedback for signal in ("rác", "chuột", "muỗi", "côn trùng", "gián")):
            return "SV-09", "IS-09-02"
        if any(signal in feedback for signal in ("mùi", "ồn", "cảnh quan", "khói", "bụi")):
            return "SV-09", "IS-09-03"
    return service_code, issue_code


def lifecycle_step_code(record: SourceRecord) -> str:
    direct_mapping = {
        "Nhận thức": "A2",
        "Xem xét": "C2",
        "Giao dịch": "TR-02",
        "Nhận nhà": "HO-03",
    }
    if record.journey_2 in direct_mapping:
        return direct_mapping[record.journey_2]
    if record.journey_2 not in {"Cư trú", "Sử dụng dịch vụ"}:
        raise ValueError(f"TC ID {record.record_key}: Journey 2 không hỗ trợ: {record.journey_2}.")
    if record.journey == "Báo sự cố":
        return "RES-07"
    if record.technical_group == "Bãi xe":
        return "RES-03"
    if record.technical_group == "App cư dân":
        return "RES-02"
    return "RES-05"


def classify(record: SourceRecord) -> ClassificationCodes:
    sentiment_mapping = {"Negative": "NEGATIVE", "Positive": "POSITIVE", "Neutral": "NEUTRAL"}
    severity_mapping = {"Cao": "SEV-2", "Trung bình": "SEV-3", "Thấp": "SEV-4"}
    try:
        sentiment = sentiment_mapping[record.sentiment]
        severity = severity_mapping[record.severity]
    except KeyError as error:
        raise ValueError(
            f"TC ID {record.record_key}: nhãn sentiment/severity không hỗ trợ."
        ) from error
    service_code, issue_code = classify_issue(record)
    step_code = lifecycle_step_code(record)
    stage_code = step_code.split("-")[0] if "-" in step_code else step_code[0]
    service_request_step_code = "SRV-02" if record.journey == "Báo sự cố" else "SRV-05"
    return ClassificationCodes(
        stage_code=stage_code,
        lifecycle_step_code=step_code,
        service_request_step_code=service_request_step_code,
        service_code=service_code,
        issue_code=issue_code,
        sentiment=sentiment,
        severity=severity,
    )


def reported_at_for(record_key: str, reference_time: datetime) -> datetime:
    """Spread source rows across 180 days when the source has no timestamp column."""
    offset = int(hashlib.sha256(record_key.encode()).hexdigest()[:8], 16) % 180
    minute = int(hashlib.sha256(f"minute:{record_key}".encode()).hexdigest()[:4], 16) % 1440
    return reference_time.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(
        days=offset, minutes=minute
    )


async def taxonomy_ids(session: Any) -> dict[str, Any]:
    release_id = await session.scalar(
        text("SELECT taxonomy_release_id FROM taxonomy_release WHERE version = :version"),
        {"version": TAXONOMY_VERSION},
    )
    if release_id is None:
        raise RuntimeError(f"Không tìm thấy taxonomy release {TAXONOMY_VERSION}.")

    async def code_map(sql: str) -> dict[str, UUID]:
        return dict((await session.execute(text(sql), {"release_id": release_id})).all())

    return {
        "release_id": release_id,
        "stage_ids": await code_map(
            "SELECT stage_code, customer_lifecycle_stage_id FROM customer_lifecycle_stage "
            "WHERE taxonomy_release_id = :release_id"
        ),
        "step_ids": await code_map(
            "SELECT step_code, customer_lifecycle_step_id FROM customer_lifecycle_step "
            "WHERE taxonomy_release_id = :release_id"
        ),
        "service_request_step_ids": await code_map(
            "SELECT step_code, service_request_step_id FROM service_request_step "
            "WHERE taxonomy_release_id = :release_id"
        ),
        "service_ids": await code_map(
            "SELECT service_code, service_id FROM service WHERE taxonomy_release_id = :release_id"
        ),
        "issue_ids": await code_map(
            "SELECT issue_code, issue_id FROM issue WHERE taxonomy_release_id = :release_id"
        ),
    }


def build_values(
    record: SourceRecord,
    codes: ClassificationCodes,
    ids: dict[str, Any],
    reference_time: datetime,
) -> dict[str, Any]:
    try:
        stage_id = ids["stage_ids"][codes.stage_code]
        lifecycle_step_id = ids["step_ids"][codes.lifecycle_step_code]
        service_request_step_id = ids["service_request_step_ids"][codes.service_request_step_code]
        service_id = ids["service_ids"][codes.service_code]
        issue_id = ids["issue_ids"][codes.issue_code]
    except KeyError as error:
        raise RuntimeError(f"Taxonomy {TAXONOMY_VERSION} thiếu mã {error.args[0]}.") from error

    return {
        "feedback_id": stable_id(f"feedback/{record.record_key}"),
        "feedback_item_id": stable_id(f"feedback-item/{record.record_key}"),
        "decision_id": stable_id(f"classification/{record.record_key}"),
        "actor_id": stable_id("actor/ocp1-loader"),
        "project_id": PROJECT_ID,
        "source_record_key": record.record_key,
        "reported_at": reported_at_for(record.record_key, reference_time),
        "content_raw": record.feedback,
        "content_masked": record.feedback,
        "raw_content_checksum": hashlib.sha256(record.feedback.encode()).hexdigest(),
        "source_metadata_json": json.dumps(
            {
                "dataset": "VOC_OCP1_100000_synthetic_unique_Journey2",
                "source_sentiment": record.sentiment,
                "source_technical_group": record.technical_group,
                "source_severity": record.severity,
                "source_journey": record.journey,
                "source_journey_2": record.journey_2,
                "nps": record.nps,
                "csat": record.csat,
                "ces": record.ces,
            },
            ensure_ascii=False,
        ),
        "taxonomy_release_id": ids["release_id"],
        "stage_id": stage_id,
        "lifecycle_step_id": lifecycle_step_id,
        "service_request_step_id": service_request_step_id,
        "service_id": service_id,
        "issue_id": issue_id,
        "sentiment": codes.sentiment,
        "severity": codes.severity,
        "decision_reason": (
            "OCP1 workbook trusted labels normalized by load_ocp1_voc_dataset.py "
            f"({record.technical_group}; {record.journey}; {record.journey_2})."
        ),
    }


DELETE_DEMO_SQL = (
    "DELETE FROM {table} WHERE {column} IN ("
    "SELECT fi.feedback_item_id FROM feedback_item fi "
    "JOIN feedback f ON f.feedback_id = fi.feedback_id "
    "WHERE f.source_system = :source_system)"
)

INSERT_FEEDBACK_SQL = text(
    """
    INSERT INTO feedback (
        feedback_id, project_id, source_system, source_record_key, reported_at,
        content_raw, content_masked, source_metadata_json, raw_content_checksum
    ) VALUES (
        :feedback_id, :project_id, 'ocp1-voc-synthetic', :source_record_key, :reported_at,
        :content_raw, :content_masked, CAST(:source_metadata_json AS json), :raw_content_checksum
    ) ON CONFLICT (source_system, source_record_key) DO NOTHING
    """
)
INSERT_ITEM_SQL = text(
    """
    INSERT INTO feedback_item (
        feedback_item_id, feedback_id, item_index, item_text_masked, status, analytic_eligibility
    ) VALUES (
        :feedback_item_id, :feedback_id, 1, :content_masked, 'ACTIVE', 'INCLUDED'
    ) ON CONFLICT (feedback_id, item_index) DO NOTHING
    """
)
INSERT_DECISION_SQL = text(
    """
    INSERT INTO classification_decision (
        classification_decision_id, feedback_item_id, decision_version, taxonomy_release_id,
        customer_lifecycle_value_status, customer_lifecycle_step_id,
        service_request_value_status, service_request_step_id,
        primary_service_value_status, primary_service_id, issue_value_status, issue_id,
        sentiment, operational_severity, cause_determination_status, classification_state,
        decision_source, decision_reason, decided_by, decided_at
    ) VALUES (
        :decision_id, :feedback_item_id, 1, :taxonomy_release_id,
        'KNOWN', :lifecycle_step_id, 'KNOWN', :service_request_step_id,
        'KNOWN', :service_id, 'KNOWN', :issue_id, :sentiment, :severity,
        'NOT_ASSESSED', 'ACCEPTED', 'SOURCE_TRUSTED', :decision_reason, :actor_id, :reported_at
    ) ON CONFLICT (feedback_item_id, decision_version) DO NOTHING
    """
)
INSERT_CURRENT_SQL = text(
    """
    INSERT INTO classification_current (
        feedback_item_id, current_decision_id, current_decision_version, taxonomy_release_id,
        customer_lifecycle_value_status, customer_lifecycle_stage_id, customer_lifecycle_step_id,
        service_request_value_status, service_request_step_id,
        primary_service_value_status, primary_service_id, issue_value_status, issue_id,
        sentiment, operational_severity, cause_determination_status, classification_state,
        last_decision_at, projection_version
    ) VALUES (
        :feedback_item_id, :decision_id, 1, :taxonomy_release_id,
        'KNOWN', :stage_id, :lifecycle_step_id, 'KNOWN', :service_request_step_id,
        'KNOWN', :service_id, 'KNOWN', :issue_id, :sentiment, :severity,
        'NOT_ASSESSED', 'ACCEPTED', :reported_at, 1
    ) ON CONFLICT (feedback_item_id) DO NOTHING
    """
)


async def delete_demo_data(session: Any) -> int:
    feedback_count = await session.scalar(
        text("SELECT count(*) FROM feedback WHERE source_system = :source_system"),
        {"source_system": DEMO_SOURCE_SYSTEM},
    )
    for table, column in (
        ("review_event", "feedback_item_id"),
        ("classification_current", "feedback_item_id"),
        ("feedback_item_hotspot", "feedback_item_id"),
        ("feedback_item_affected_channel", "feedback_item_id"),
    ):
        await session.execute(
            text(DELETE_DEMO_SQL.format(table=table, column=column)),
            {"source_system": DEMO_SOURCE_SYSTEM},
        )
    await session.execute(
        text(
            "DELETE FROM classification_decision WHERE feedback_item_id IN ("
            "SELECT fi.feedback_item_id FROM feedback_item fi JOIN feedback f "
            "ON f.feedback_id = fi.feedback_id WHERE f.source_system = :source_system)"
        ),
        {"source_system": DEMO_SOURCE_SYSTEM},
    )
    await session.execute(
        text(
            "DELETE FROM feedback_item WHERE feedback_id IN ("
            "SELECT feedback_id FROM feedback WHERE source_system = :source_system)"
        ),
        {"source_system": DEMO_SOURCE_SYSTEM},
    )
    await session.execute(
        text("DELETE FROM feedback WHERE source_system = :source_system"),
        {"source_system": DEMO_SOURCE_SYSTEM},
    )
    return int(feedback_count or 0)


def chunks(records: list[dict[str, Any]], size: int) -> Iterator[list[dict[str, Any]]]:
    for index in range(0, len(records), size):
        yield records[index : index + size]


async def load_records(records: list[SourceRecord], batch_size: int) -> tuple[int, Counter[str]]:
    reference_time = datetime.now(timezone.utc)
    classified = [(record, classify(record)) for record in records]
    distributions = Counter(f"{codes.service_code}/{codes.issue_code}" for _, codes in classified)

    async with AsyncSessionLocal() as session:
        async with session.begin():
            ids = await taxonomy_ids(session)
            values = [build_values(record, codes, ids, reference_time) for record, codes in classified]
            deleted_demo_count = await delete_demo_data(session)
            for batch in chunks(values, batch_size):
                await session.execute(INSERT_FEEDBACK_SQL, batch)
                await session.execute(INSERT_ITEM_SQL, batch)
                await session.execute(INSERT_DECISION_SQL, batch)
                await session.execute(INSERT_CURRENT_SQL, batch)
    return deleted_demo_count, distributions


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="Path to the OCP1 .xlsx workbook.")
    parser.add_argument("--batch-size", type=int, default=1_000, help="Rows inserted per SQL batch.")
    parser.add_argument("--dry-run", action="store_true", help="Validate and summarize only; do not write PostgreSQL.")
    return parser.parse_args()


async def main() -> None:
    args = parse_args()
    if args.batch_size < 1:
        raise ValueError("--batch-size phải lớn hơn 0.")
    if not args.input.is_file():
        raise FileNotFoundError(f"Không tìm thấy workbook: {args.input}")

    records = load_source_records(args.input)
    classified = [(record, classify(record)) for record in records]
    print(f"Validated {len(records):,} workbook rows from {args.input.name}.")
    for title, values in (
        ("Sentiment", Counter(codes.sentiment for _, codes in classified)),
        ("Severity", Counter(codes.severity for _, codes in classified)),
        ("Journey steps", Counter(codes.lifecycle_step_code for _, codes in classified)),
        ("Issues", Counter(f"{codes.service_code}/{codes.issue_code}" for _, codes in classified)),
    ):
        print(f"{title}: {dict(sorted(values.items()))}")
    if args.dry_run:
        return

    try:
        deleted_demo_count, distributions = await load_records(records, args.batch_size)
        print(f"Deleted {deleted_demo_count} analytics-demo feedback records.")
        print(f"Loaded/classified {len(records):,} OCP1 feedback records into taxonomy {TAXONOMY_VERSION}.")
        print(f"Issue distribution: {dict(sorted(distributions.items()))}")
    finally:
        await engine.dispose()


if __name__ == "__main__":
    asyncio.run(main())
